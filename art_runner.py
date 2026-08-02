#!/usr/bin/env python3
"""
art_runner.py — ART regression run: execute a report against the pre and
post databases and diff the results.

Usage:
    python art_runner.py <report_name> [--iterations N]

Flow:
  1. Read the report's metadata from the ART tables on the PRE database
     (art.reports_metadata / art.params_metadata / art.candidates_<report>).
  2. Derive the unresolved param space automatically:
       - each XOR pair of business keys (p_entity_id vs p_group_id)
         contributes two branches: one side supplied, the other NULL
       - each enum param contributes one branch per declared value
         (the literal value 'NULL' means SQL NULL)
       - mandatory unpaired business keys appear in every branch
     The space is the cartesian product of those dimensions.
  3. Resolve each combination against the first N candidates
     (candidate_id order — deterministic; N = --iterations, default all).
     One (combination x candidate) = one execution.
  4. Run every execution identically on PRE and POST.
  5. Compare per execution, keyed on the report's declared primary key
     (pk_nullable columns join NULL-safely):
       - pk present on one side only  -> row_missing_in_pre / row_missing_in_post
       - pk on both sides, any non-PK column differs -> value_mismatch
         (exact by default; reports_metadata.tolerances can declare
          'col=eps;col=eps' for per-column numeric tolerance)
       - duplicated pk within one side -> duplicate_pk (comparison for
         that key is skipped; the duplication itself is the finding)
  6. Export outputs/<report>_<timestamp>.xlsx with tabs:
       Pre_output, Post_output   — all rows from every execution
       Reports, Params, Candidates — the metadata that drove the run
       Flagged_Diffs             — one row per finding, incl. the exact
                                   resolved query to paste-and-run
       Summary                   — pass/fail + counts per execution,
                                   distinct diff columns, coverage notes

Exit code: 0 = ran and no diffs; 1 = ran and diffs found; 2 = setup error.
"""

import argparse
import os
import sys
from datetime import datetime, date
from decimal import Decimal, InvalidOperation
from itertools import product
from pathlib import Path

import psycopg2
from dotenv import load_dotenv

PROJECT_DIR = Path(__file__).parent
OUTPUT_DIR = PROJECT_DIR / "outputs"

load_dotenv(PROJECT_DIR / ".env")

DB_HOST = os.environ.get("DB_HOST", "localhost")
DB_PORT = int(os.environ.get("DB_PORT", "5432"))
DB_USER = os.environ.get("DB_USER")
DB_PASSWORD = os.environ.get("DB_PASSWORD")
PRE_DB = os.environ.get("PRE_DB_NAME", "art_pre")
POST_DB = os.environ.get("POST_DB_NAME", "art_post")
ART_SCHEMA = os.environ.get("ART_SCHEMA", "art").strip().lower()


def connect(dbname):
    return psycopg2.connect(host=DB_HOST, port=DB_PORT, user=DB_USER,
                            password=DB_PASSWORD, dbname=dbname)


# --------------------------------------------------------------------
# Metadata loading (PRE is the source of truth)
# --------------------------------------------------------------------
def load_metadata(conn, report_name):
    with conn.cursor() as cur:
        cur.execute(
            f"SELECT report_id, report_type, target, primary_key, pk_nullable, "
            f"candidates_table, tolerances FROM {ART_SCHEMA}.reports_metadata "
            f"WHERE report_name = %s", (report_name,))
        row = cur.fetchone()
        if not row:
            sys.exit(f"Report '{report_name}' is not onboarded — run art_setup_config.py first. "
                     f"(checked {ART_SCHEMA}.reports_metadata on {PRE_DB})")
        report = {
            "report_id": row[0], "report_type": row[1], "target": row[2],
            "primary_key": [c.strip() for c in row[3].split(",")],
            "pk_nullable": [c.strip() for c in row[4].split(",")] if row[4] else [],
            "candidates_table": row[5],
            "tolerances": parse_tolerances(row[6]),
        }
        if report["report_type"] != "sql_function":
            sys.exit(f"report_type '{report['report_type']}' not supported by this runner "
                     f"(sql_function only today).")

        cur.execute(
            f"SELECT param_id, param_name, param_kind, enum_values, optional_pair, nullable "
            f"FROM {ART_SCHEMA}.params_metadata WHERE report_id = %s ORDER BY param_id",
            (report["report_id"],))
        params = [
            {"param_id": r[0], "name": r[1], "kind": r[2],
             "enum_values": [v.strip() for v in r[3].split(",")] if r[3] else None,
             "optional_pair": r[4], "nullable": r[5]}
            for r in cur.fetchall()
        ]

        candidates = []
        if report["candidates_table"]:
            # The metadata row can outlive its candidates table if onboarding
            # was interrupted between the two writes. Check rather than let
            # psycopg2 raise UndefinedTable at the user.
            cur.execute("SELECT to_regclass(%s)",
                        (f"{ART_SCHEMA}.{report['candidates_table']}",))
            if cur.fetchone()[0] is None:
                sys.exit(
                    f"Report '{report_name}' is registered but its candidates table "
                    f"{ART_SCHEMA}.{report['candidates_table']} does not exist — "
                    f"onboarding was interrupted before it was created. Re-run "
                    f"art_setup_config.py for this report.")
            cur.execute(f"SELECT * FROM {ART_SCHEMA}.{report['candidates_table']} "
                        f"ORDER BY candidate_id")
            cand_cols = [d[0] for d in cur.description]
            candidates = [dict(zip(cand_cols, r)) for r in cur.fetchall()]

    return report, params, candidates


def parse_tolerances(raw):
    """'col=eps;col=eps' -> {col: Decimal(eps)}. NULL/empty -> {}."""
    if not raw:
        return {}
    out = {}
    for part in raw.split(";"):
        part = part.strip()
        if not part:
            continue
        col, _, eps = part.partition("=")
        out[col.strip()] = Decimal(eps.strip())
    return out


# --------------------------------------------------------------------
# Unresolved param space derivation
# --------------------------------------------------------------------
def derive_param_space(params):
    """Returns a list of 'combos'. Each combo is a dict param_name ->
    spec, where spec is one of:
        ("candidate", name)  resolve from the candidate row's column
        ("literal", value)   fixed value for this branch (enum branches;
                             value None means SQL NULL)
    Dimensions:
      - XOR pair (a, b): two branches — a from candidate + b NULL, or
        b from candidate + a NULL
      - nullable business key: two branches — from candidate, or NULL
        (modeled per the metadata schema; no demo report sets nullable)
      - enum: one branch per declared value ('NULL' -> SQL NULL)
      - mandatory unpaired business key: always from the candidate
    """
    fixed = {}          # param -> spec present in every combo
    dimensions = []     # list of lists of {param: spec} alternatives

    seen_pair = set()
    for p in params:
        name = p["name"]
        if p["kind"] == "enum":
            dimensions.append([
                {name: ("literal", None if v == "NULL" else v)}
                for v in p["enum_values"]
            ])
        elif p["optional_pair"]:
            if name in seen_pair:
                continue
            partner = p["optional_pair"]
            seen_pair.update((name, partner))
            dimensions.append([
                {name: ("candidate", name), partner: ("literal", None)},
                {partner: ("candidate", partner), name: ("literal", None)},
            ])
        elif p["nullable"]:
            dimensions.append([
                {name: ("candidate", name)},
                {name: ("literal", None)},
            ])
        else:
            fixed[name] = ("candidate", name)

    combos = []
    for parts in product(*dimensions) if dimensions else [()]:
        combo = dict(fixed)
        for part in parts:
            combo.update(part)
        combos.append(combo)
    return combos


def resolve_combo(combo, params, candidate):
    """-> ordered list of values matching the routine's signature order,
    plus a human-readable resolved-params string."""
    values, described = [], []
    for p in params:
        spec_kind, spec_val = combo[p["name"]]
        if spec_kind == "candidate":
            if spec_val not in candidate:
                sys.exit(f"Candidate row {candidate.get('candidate_id')} has no column "
                         f"'{spec_val}' — candidates table out of sync with params metadata.")
            v = candidate[spec_val]
        else:
            v = spec_val
        values.append(v)
        described.append(f"{p['name']}={sql_literal(v)}")
    return values, ", ".join(described)


def sql_literal(v):
    """Render a Python value as a SQL literal for the paste-and-run
    debug query. Executions themselves use parameterized calls — this
    is display only."""
    if v is None:
        return "NULL"
    if isinstance(v, bool):
        return "TRUE" if v else "FALSE"
    if isinstance(v, (int, float, Decimal)):
        return str(v)
    if isinstance(v, (date, datetime)):
        return f"'{v.isoformat()}'"
    return "'" + str(v).replace("'", "''") + "'"


# --------------------------------------------------------------------
# Execution
# --------------------------------------------------------------------
def execute_report(conn, target, values):
    placeholders = ", ".join(["%s"] * len(values))
    with conn.cursor() as cur:
        cur.execute(f"SELECT * FROM {target}({placeholders})", values)
        columns = [d[0] for d in cur.description]
        rows = cur.fetchall()
    return columns, rows


# --------------------------------------------------------------------
# Comparison
# --------------------------------------------------------------------
def index_by_pk(columns, rows, primary_key):
    """-> ({pk_tuple: row}, [duplicated pk_tuples]). NULLs inside the
    tuple compare NULL-safely for free (None == None in a dict key)."""
    pk_idx = [columns.index(c) for c in primary_key]
    indexed, dups = {}, []
    for row in rows:
        key = tuple(row[i] for i in pk_idx)
        if key in indexed:
            dups.append(key)
        else:
            indexed[key] = row
    return indexed, dups


def values_differ(col, a, b, tolerances):
    if a is None and b is None:
        return False
    if (a is None) != (b is None):
        return True
    eps = tolerances.get(col)
    if eps is not None:
        try:
            return abs(Decimal(str(a)) - Decimal(str(b))) > eps
        except (InvalidOperation, TypeError):
            pass  # not numeric after all — fall through to exact
    return a != b


def compare_execution(exec_meta, columns, pre_rows, post_rows, primary_key, tolerances):
    """-> list of diff dicts for one execution."""
    diffs = []
    pre_idx, pre_dups = index_by_pk(columns, pre_rows, primary_key)
    post_idx, post_dups = index_by_pk(columns, post_rows, primary_key)

    def flag(diff_type, pk, column=None, pre_v=None, post_v=None):
        diffs.append({**exec_meta, "diff_type": diff_type,
                      "pk": pk, "column": column,
                      "pre_value": pre_v, "post_value": post_v})

    skipped = set()
    for key in set(pre_dups) | set(post_dups):
        flag("duplicate_pk", key)
        skipped.add(key)

    for key in pre_idx.keys() - post_idx.keys():
        if key not in skipped:
            flag("row_missing_in_post", key)
    for key in post_idx.keys() - pre_idx.keys():
        if key not in skipped:
            flag("row_missing_in_pre", key)

    non_pk = [c for c in columns if c not in primary_key]
    for key in pre_idx.keys() & post_idx.keys():
        if key in skipped:
            continue
        pre_row, post_row = pre_idx[key], post_idx[key]
        for col in non_pk:
            i = columns.index(col)
            if values_differ(col, pre_row[i], post_row[i], tolerances):
                flag("value_mismatch", key, col, pre_row[i], post_row[i])
    return diffs


# --------------------------------------------------------------------
# xlsx export
# --------------------------------------------------------------------
def export_xlsx(report_name, report, params, candidates, executions, diffs, path):
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    FONT = "Arial"
    base = Font(name=FONT, size=10)
    bold = Font(name=FONT, size=10, bold=True)
    fills = {
        "value_mismatch":      PatternFill("solid", start_color="FFF2CC"),
        "row_missing_in_post": PatternFill("solid", start_color="F8CBAD"),
        "row_missing_in_pre":  PatternFill("solid", start_color="F8CBAD"),
        "duplicate_pk":        PatternFill("solid", start_color="D9D2E9"),
    }

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    def sheet(title, header, rows, fill_by=None):
        ws = wb.create_sheet(title)
        ws.append(header)
        for c in ws[1]:
            c.font = bold
        for r in rows:
            # native types (numbers, dates) go through untouched so Excel
            # sorts and filters them properly; everything else is stringified
            ws.append([cell if cell is None or isinstance(
                cell, (int, float, Decimal, date, datetime))
                else str(cell) for cell in r])
            for c in ws[ws.max_row]:
                if isinstance(c.value, datetime):
                    c.number_format = "yyyy-mm-dd hh:mm:ss"
                elif isinstance(c.value, date):
                    c.number_format = "yyyy-mm-dd"
            if fill_by:
                f = fills.get(fill_by(r))
                if f:
                    for c in ws[ws.max_row]:
                        c.fill = f
        for row in ws.iter_rows(min_row=2):
            for c in row:
                c.font = base
        ws.freeze_panes = "A2"
        for i, col in enumerate(header, start=1):
            width = max(len(str(col)) + 2, 12)
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = min(width, 40)
        return ws

    # Pre_output / Post_output — every row from every execution, tagged
    out_header = ["execution_id", "candidate_id", "resolved_params"] + executions[0]["columns"]
    for tab, side in (("Pre_output", "pre_rows"), ("Post_output", "post_rows")):
        rows = []
        for ex in executions:
            for r in ex[side]:
                rows.append([ex["execution_id"], ex["candidate_id"], ex["resolved_params"], *r])
        sheet(tab, out_header, rows)

    # Metadata tabs — what actually drove the run
    sheet("Reports",
          ["report_name", "report_type", "target", "primary_key", "pk_nullable",
           "candidates_table", "tolerances"],
          [[report_name, report["report_type"], report["target"],
            ",".join(report["primary_key"]),
            ",".join(report["pk_nullable"]) or None,
            report["candidates_table"],
            ";".join(f"{k}={v}" for k, v in report["tolerances"].items()) or None]])
    sheet("Params",
          ["param_id", "param_name", "param_kind", "enum_values", "optional_pair", "nullable"],
          [[p["param_id"], p["name"], p["kind"],
            ",".join(p["enum_values"]) if p["enum_values"] else None,
            p["optional_pair"], p["nullable"]] for p in params])
    if candidates:
        cand_header = list(candidates[0].keys())
        sheet("Candidates", cand_header, [[c[k] for k in cand_header] for c in candidates])

    # Flagged_Diffs — one row per finding + the paste-and-run query
    sheet("Flagged_Diffs",
          ["execution_id", "candidate_id", "resolved_params", "diff_type",
           "pk_values", "column", "pre_value", "post_value", "resolved_query"],
          [[d["execution_id"], d["candidate_id"], d["resolved_params"], d["diff_type"],
            ", ".join(sql_literal(v) for v in d["pk"]), d["column"],
            d["pre_value"], d["post_value"], d["resolved_query"]] for d in diffs],
          fill_by=lambda r: r[3])

    # Summary
    by_exec = {}
    for d in diffs:
        by_exec.setdefault(d["execution_id"], []).append(d)
    summary_rows = []
    for ex in executions:
        ds = by_exec.get(ex["execution_id"], [])
        counts = {}
        for d in ds:
            counts[d["diff_type"]] = counts.get(d["diff_type"], 0) + 1
        summary_rows.append([
            ex["execution_id"], ex["candidate_id"], ex["resolved_params"],
            len(ex["pre_rows"]), len(ex["post_rows"]),
            "FAIL" if ds else "PASS",
            counts.get("value_mismatch", 0),
            counts.get("row_missing_in_pre", 0),
            counts.get("row_missing_in_post", 0),
            counts.get("duplicate_pk", 0),
        ])
    ws = sheet("Summary",
               ["execution_id", "candidate_id", "resolved_params", "pre_rows", "post_rows",
                "result", "value_mismatches", "rows_missing_in_pre", "rows_missing_in_post",
                "duplicate_pks"],
               summary_rows,
               fill_by=lambda r: "row_missing_in_post" if r[5] == "FAIL" else None)

    # distinct flagged columns + totals under the per-execution table
    ws.append([])
    distinct_cols = sorted({d["column"] for d in diffs if d["column"]})
    totals = {}
    for d in diffs:
        totals[d["diff_type"]] = totals.get(d["diff_type"], 0) + 1
    from openpyxl.styles import Font as F
    meta_lines = [
        ["distinct columns with value mismatches:", ", ".join(distinct_cols) or "none"],
        ["total diffs by type:",
         ", ".join(f"{k}={v}" for k, v in sorted(totals.items())) or "none"],
        ["executions:", len(executions)],
        ["overall:", "FAIL" if diffs else "PASS"],
    ]
    for line in meta_lines:
        ws.append(line)
        ws[ws.max_row][0].font = F(name=FONT, size=10, bold=True)
        for c in ws[ws.max_row]:
            if not c.font.bold:
                c.font = F(name=FONT, size=10)

    OUTPUT_DIR.mkdir(exist_ok=True)
    wb.save(path)


# --------------------------------------------------------------------
# Main
# --------------------------------------------------------------------
def main():
    ap = argparse.ArgumentParser(description="Run an ART pre/post regression for a report.")
    ap.add_argument("report_name")
    ap.add_argument("--iterations", type=int, default=None,
                    help="number of candidates to use, first-N by candidate_id "
                         "(default: all)")
    args = ap.parse_args()

    if not DB_USER or not DB_PASSWORD:
        sys.exit("DB_USER / DB_PASSWORD missing — fill in .env first.")

    pre_conn, post_conn = connect(PRE_DB), connect(POST_DB)
    try:
        report, params, candidates = load_metadata(pre_conn, args.report_name)

        combos = derive_param_space(params)
        needs_candidates = any(spec[0] == "candidate"
                               for combo in combos for spec in combo.values())
        if needs_candidates and not candidates:
            sys.exit(f"No candidate rows in {ART_SCHEMA}.{report['candidates_table']} — "
                     f"fill them in (art_setup_config.py gates on this).")

        used = candidates
        coverage_note = None
        if args.iterations is not None:
            if args.iterations < len(candidates):
                used = candidates[:args.iterations]
            elif args.iterations > len(candidates):
                coverage_note = (f"iterations={args.iterations} requested but only "
                                 f"{len(candidates)} candidate(s) exist — using all")
        if not needs_candidates:
            used = [{"candidate_id": None}]   # combos are fully literal

        print(f"Report '{args.report_name}' ({report['target']}) — "
              f"{len(combos)} unresolved combination(s) x {len(used)} candidate(s) "
              f"= {len(combos) * len(used)} execution(s) per DB")
        if coverage_note:
            print(f"  note: {coverage_note}")

        executions, all_diffs = [], []
        exec_id = 0
        for candidate in used:
            for combo in combos:
                exec_id += 1
                values, described = resolve_combo(combo, params, candidate)
                query = (f"SELECT * FROM {report['target']}"
                         f"({', '.join(sql_literal(v) for v in values)});")

                pre_cols, pre_rows = execute_report(pre_conn, report["target"], values)
                post_cols, post_rows = execute_report(post_conn, report["target"], values)
                if pre_cols != post_cols:
                    sys.exit(f"Column sets differ between pre and post for execution "
                             f"{exec_id} — schemas are out of sync: {pre_cols} vs {post_cols}")

                ex = {"execution_id": exec_id,
                      "candidate_id": candidate["candidate_id"],
                      "resolved_params": described,
                      "resolved_query": query,
                      "columns": pre_cols,
                      "pre_rows": pre_rows, "post_rows": post_rows}
                executions.append(ex)

                diffs = compare_execution(
                    {"execution_id": exec_id, "candidate_id": candidate["candidate_id"],
                     "resolved_params": described, "resolved_query": query},
                    pre_cols, pre_rows, post_rows,
                    report["primary_key"], report["tolerances"])
                all_diffs.extend(diffs)

                status = f"{len(diffs)} diff(s)" if diffs else "clean"
                print(f"  [{exec_id:>3}] cand={candidate['candidate_id']}  "
                      f"{described}  ->  pre={len(pre_rows)} post={len(post_rows)} rows, {status}")
    finally:
        pre_conn.close()
        post_conn.close()

    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = OUTPUT_DIR / f"{args.report_name}_{stamp}.xlsx"
    export_xlsx(args.report_name, report, params, candidates, executions, all_diffs, out_path)

    total = len(all_diffs)
    print(f"\n{'FAIL' if total else 'PASS'} — {total} diff(s) across "
          f"{len(executions)} execution(s). Report: {out_path}")
    sys.exit(1 if total else 0)


if __name__ == "__main__":
    main()
